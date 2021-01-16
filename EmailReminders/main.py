import openpyxl, smtplib

# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('members.xlsx')
sheet = wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email  # dictionary: key-value pairs

# TODO: Log in to email account.
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)  # smtp.office365.com  smtp.mail.yahoo.com

smtpObj.ehlo()

smtpObj.starttls()
smtpObj.login('your email', 'your password')
# sys.argv[1])

# TODO: Send out reminder emails.

# Send out reminder emails.
for name, email in unpaidMembers.items():
    # String interpolation
    body = ('Subject: %s dues unpaid.\n\nDear %s,\nRecords show that you have not paid dues for %s. '
            'Please make this payment as soon as possible. Thank you!' % (latestMonth, name, latestMonth))
    print('Sending email to %s...' % email)

    sendmailStatus = smtpObj.sendmail('sender email', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()
