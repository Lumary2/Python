import openpyxl

wb= openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)

sheet=wb['Sheet1']
print(sheet.title)

for i in range(1, sheet.max_row+1):
    print(sheet.cell(row=i,column=2).value)

sheet['B1']='Cat'

wb.save('example.xlsx')